from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select

from app.models import Blacklist, Category, Product, async_session


async def build_products_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Артикул", "Название", "Бренд", "Категория",
        "Цена дилера", "Цена розн.", "Цена OMarket",
        "Наценка %", "Остаток", "Активен",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3B82F6")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    async with async_session() as session:
        cat_rows = (await session.execute(select(Category.id, Category.name))).fetchall()
        cat_names = {c[0]: c[1] for c in cat_rows}

        blacklisted = {
            r[0] for r in (await session.execute(select(Blacklist.article))).fetchall()
        }

        products = (await session.execute(
            select(Product).order_by(Product.article)
        )).scalars().all()

        for p in products:
            markup_pct = None
            if p.price_dealer and p.price_omarket and p.price_dealer > 0:
                markup_pct = round((p.price_omarket / p.price_dealer - 1) * 100, 1)
            ws.append([
                p.article,
                p.name,
                p.brand or "",
                cat_names.get(p.category_id, ""),
                p.price_dealer or 0,
                p.price_retail or 0,
                p.price_omarket or 0,
                markup_pct if markup_pct is not None else "",
                p.quantity or "",
                "Нет" if p.article in blacklisted else ("Да" if p.is_active else "Нет"),
            ])

    widths = [10, 50, 16, 30, 14, 14, 14, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
